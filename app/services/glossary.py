import csv
import uuid
import os
import re
from openpyxl import load_workbook
from typing import Optional

class GlossaryService:
    def __init__(self):
        self._glossaries: dict[str, dict[str, str]] = {}
        self._metadata: dict[str, dict] = {}

    def load_glossary(self, file_path: str, filename: str) -> str:
        glossary_id = str(uuid.uuid4())
        terms = {}
        ext = os.path.splitext(filename)[-1].lower()
        if ext == ".csv":
            terms = self._load_csv(file_path)
        elif ext in [".xlsx"]:  # openpyxl does not support .xls
            terms = self._load_xlsx(file_path)
        else:
            raise ValueError(f"Unsupported glossary format: {ext}")
        self._glossaries[glossary_id] = terms
        self._metadata[glossary_id] = {
            "filename": filename,
            "term_count": len(terms)
        }
        return glossary_id

    def _load_csv(self, path: str) -> dict[str, str]:
        terms = {}
        with open(path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            for row in reader:
                if len(row) >= 2:
                    cn, en = row[0].strip(), row[1].strip()
                    if cn and en and cn != "中文术语":
                        terms[cn] = en
        return terms

    def _load_xlsx(self, path: str) -> dict[str, str]:
        terms = {}
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            if len(row) >= 2:
                if row[0] is not None and row[1] is not None:
                    cn, en = str(row[0]).strip(), str(row[1]).strip()
                    if cn and en and cn != "中文术语":
                        terms[cn] = en
        return terms

    def get_glossary(self, glossary_id: str) -> dict[str, str]:
        if glossary_id not in self._glossaries:
            raise ValueError(f"Glossary not found: {glossary_id}")
        return self._glossaries[glossary_id]

    def get_term_count(self, glossary_id: str) -> int:
        if glossary_id not in self._glossaries:
            raise ValueError(f"Glossary not found: {glossary_id}")
        return len(self._glossaries[glossary_id])

    def get_metadata(self, glossary_id: str) -> Optional[dict]:
        if glossary_id not in self._metadata:
            raise ValueError(f"Glossary not found: {glossary_id}")
        return self._metadata[glossary_id]

    def normalize_quotes(self, text: str) -> str:
        """Normalize straight double quotes to curly double quotes. Idempotent."""
        # Step 1: normalize any existing curly quotes back to straight (ensures idempotency)
        result = text.replace(chr(0x201C), chr(34)).replace(chr(0x201D), chr(34))
        # Step 2: pair straight quotes — odd positions become left, even become right
        parts = result.split(chr(34))
        result = ""
        for i, part in enumerate(parts):
            if i == len(parts) - 1:
                result += part
            elif i % 2 == 0:
                result += part + chr(0x201C)
            else:
                result += part + chr(0x201D)
        return result

    def replace_with_glossary(self, text: str, glossary: dict[str, str]) -> str:
        """Longest-match-first replacement using glossary dict."""
        if not text or not glossary:
            return text
        # Normalize quotes in text first
        text = self.normalize_quotes(text)
        # Normalize glossary keys for matching
        normalized_glossary = {self.normalize_quotes(k): v for k, v in glossary.items()}
        sorted_terms = sorted(normalized_glossary.keys(), key=len, reverse=True)
        result = text
        for term in sorted_terms:
            pattern = re.escape(term)
            result = re.sub(pattern, normalized_glossary[term], result)
        return result
